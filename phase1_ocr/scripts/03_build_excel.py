#!/usr/bin/env python3
"""Build review Excel from OCR JSON files.

Output sheet structure (one row per session):
| 來源頁 | 日期 | 時段 | 時刻 | 氣溫 |
| L1收 L1舒 L1脈 | R1收 R1舒 R1脈 | L2收 L2舒 L2脈 | R2收 R2舒 R2脈 |
| 備註 | 信心度 | OCR原文 | 校對狀態 | 校對備註 |

Visual aids:
- Alternate light grey/white shading per day
- Low confidence row → light red background on confidence cell
- Out-of-range BP cells → yellow background
- "uncertain" cells → orange background
- 來源頁 cell links to the page PNG
"""
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink

ROOT = Path(__file__).resolve().parents[1]
JSON_DIR = ROOT / "ocr_raw"
OUT_DIR = ROOT / "output"
PAGES_DIR = ROOT / "pages"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_FILE = OUT_DIR / "bp_ocr_review.xlsx"

HEADERS = [
    "來源頁", "日期", "時段", "時刻", "氣溫°C",
    "L1收", "L1舒", "L1脈",
    "R1收", "R1舒", "R1脈",
    "L2收", "L2舒", "L2脈",
    "R2收", "R2舒", "R2脈",
    "備註", "信心度", "OCR原文", "校對狀態", "校對備註",
]

# Range checks
BP_RANGES = {
    "systolic": (70, 250),
    "diastolic": (40, 150),
    "pulse": (30, 200),
}

# Styles
fill_grey = PatternFill("solid", fgColor="F2F2F2")
fill_white = PatternFill("solid", fgColor="FFFFFF")
fill_low_conf = PatternFill("solid", fgColor="FFCCCC")
fill_med_conf = PatternFill("solid", fgColor="FFF4CC")
fill_oor = PatternFill("solid", fgColor="FFE699")  # out-of-range yellow
fill_uncertain = PatternFill("solid", fgColor="FFB266")  # orange
fill_dup = PatternFill("solid", fgColor="FF6666")  # duplicate (date, period) — red
fill_header = PatternFill("solid", fgColor="305496")
font_header = Font(bold=True, color="FFFFFF", size=11)
border_thin = Border(*(Side(style="thin", color="BFBFBF"),) * 4)


def load_sessions():
    sessions = []
    for jf in sorted(JSON_DIR.glob("page_*.json")):
        page_num = int(jf.stem.split("_")[1])
        try:
            data = json.loads(jf.read_text())
        except json.JSONDecodeError as e:
            print(f"WARN: {jf.name} parse error: {e}")
            continue
        for sess in data.get("sessions", []):
            sess["_page"] = page_num
            sessions.append(sess)
    sessions.sort(key=lambda s: (s.get("date", ""), 0 if s.get("period") == "AM" else 1))
    return sessions


def find_reading(readings, seq, arm):
    for r in readings or []:
        if r.get("seq") == seq and r.get("arm") == arm:
            return r
    return None


def field_is_uncertain(reading, field):
    if not reading:
        return False
    return field in (reading.get("uncertain") or [])


def field_out_of_range(reading, field):
    if not reading:
        return False
    v = reading.get(field)
    if v is None:
        return False
    lo, hi = BP_RANGES[field]
    return v < lo or v > hi


def build():
    sessions = load_sessions()
    print(f"Loaded {len(sessions)} sessions from {len(list(JSON_DIR.glob('page_*.json')))} pages")

    # Find duplicate (date, period) for highlighting
    from collections import defaultdict
    sess_keys = defaultdict(list)
    for i, s in enumerate(sessions):
        sess_keys[(s.get("date"), s.get("period"))].append(i)
    dup_indices = set()
    for k, idxs in sess_keys.items():
        if len(idxs) > 1:
            dup_indices.update(idxs)

    wb = Workbook()
    ws = wb.active
    ws.title = "校對檢視"
    ws.freeze_panes = "B2"

    # Header row
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill_header
        c.font = font_header
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border_thin

    # Data validation for 校對狀態 column
    status_col_idx = HEADERS.index("校對狀態") + 1
    dv = DataValidation(type="list", formula1='"未校對,已確認,已修正,有問題"', allow_blank=True)
    ws.add_data_validation(dv)

    last_date = None
    use_grey = False
    fields = ["systolic", "diastolic", "pulse"]

    for i, sess in enumerate(sessions, start=2):
        sess_idx = i - 2
        date = sess.get("date", "")
        if date != last_date:
            use_grey = not use_grey
            last_date = date
        is_dup = sess_idx in dup_indices
        row_fill = fill_dup if is_dup else (fill_grey if use_grey else fill_white)

        # Source page hyperlink
        page_num = sess.get("_page")
        page_name = f"p{page_num:02d}"
        cell = ws.cell(row=i, column=1, value=page_name)
        png_path = PAGES_DIR / f"page_{page_num:02d}.png"
        if png_path.exists():
            cell.hyperlink = Hyperlink(ref=cell.coordinate, target=str(png_path.absolute()))
            cell.font = Font(color="0563C1", underline="single")

        ws.cell(row=i, column=2, value=date)
        ws.cell(row=i, column=3, value=sess.get("period", ""))
        ws.cell(row=i, column=4, value=sess.get("time", ""))
        ws.cell(row=i, column=5, value=sess.get("temperature_c"))

        # 4 readings × 3 fields = 12 BP cells (cols 6..17)
        col = 6
        for seq, arm in [(1, "L"), (1, "R"), (2, "L"), (2, "R")]:
            r = find_reading(sess.get("readings", []), seq, arm)
            for field in fields:
                v = r.get(field) if r else None
                cell = ws.cell(row=i, column=col, value=v)
                if field_is_uncertain(r, field):
                    cell.fill = fill_uncertain
                elif field_out_of_range(r, field):
                    cell.fill = fill_oor
                col += 1

        ws.cell(row=i, column=18, value=sess.get("notes", ""))
        conf = sess.get("confidence", "")
        cc = ws.cell(row=i, column=19, value=conf)
        if conf == "low":
            cc.fill = fill_low_conf
        elif conf == "medium":
            cc.fill = fill_med_conf
        ws.cell(row=i, column=20, value=sess.get("raw_text", ""))
        sc = ws.cell(row=i, column=21, value="未校對")
        dv.add(sc)
        ws.cell(row=i, column=22, value="")

        # Apply row fill to non-highlighted cells
        for col_i in range(1, len(HEADERS) + 1):
            c = ws.cell(row=i, column=col_i)
            if c.fill.fgColor.rgb in ("00FFFFFF", "FFFFFFFF", None):
                c.fill = row_fill
            c.border = border_thin
            c.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    widths = {
        "A": 7, "B": 12, "C": 6, "D": 8, "E": 8,
        "F": 7, "G": 7, "H": 7, "I": 7, "J": 7, "K": 7,
        "L": 7, "M": 7, "N": 7, "O": 7, "P": 7, "Q": 7,
        "R": 24, "S": 8, "T": 36, "U": 12, "V": 24,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 32

    # Summary sheet
    ws2 = wb.create_sheet("統計")
    ws2["A1"] = "Phase 1 OCR 統計"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A3"] = "頁數"
    ws2["B3"] = len(set(s["_page"] for s in sessions))
    ws2["A4"] = "Session 數"
    ws2["B4"] = len(sessions)
    ws2["A5"] = "完整 4 筆讀數的 session"
    ws2["B5"] = sum(1 for s in sessions if len(s.get("readings", [])) == 4)
    ws2["A6"] = "信心度 high"
    ws2["B6"] = sum(1 for s in sessions if s.get("confidence") == "high")
    ws2["A7"] = "信心度 medium"
    ws2["B7"] = sum(1 for s in sessions if s.get("confidence") == "medium")
    ws2["A8"] = "信心度 low"
    ws2["B8"] = sum(1 for s in sessions if s.get("confidence") == "low")
    ws2["A9"] = "uncertain 欄位數"
    ws2["B9"] = sum(
        len(r.get("uncertain", []))
        for s in sessions for r in s.get("readings", [])
    )
    ws2["A10"] = "重複 (date, period) session 數"
    ws2["B10"] = len(dup_indices)
    dates = sorted(set(s.get("date","") for s in sessions if s.get("date")))
    ws2["A11"] = "日期範圍"
    ws2["B11"] = f"{dates[0]} ~ {dates[-1]}" if dates else "?"
    ws2["A12"] = "唯一日期數"
    ws2["B12"] = len(dates)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 22

    # Duplicate report sheet
    ws_dup = wb.create_sheet("重複日期")
    ws_dup["A1"] = "重複 (日期+時段) 詳細列表"
    ws_dup["A1"].font = Font(bold=True, size=14)
    ws_dup["A2"] = "同一個 (日期, 時段) 出現在多頁,代表 OCR 把某一頁的日期讀錯。請點來源頁超連結比對原圖,把錯誤的列在主分頁標『有問題』。"
    ws_dup["A2"].alignment = Alignment(wrap_text=True)
    ws_dup.row_dimensions[2].height = 50
    ws_dup.merge_cells("A2:H2")

    dup_headers = ["日期", "時段", "出現頁", "L1收/舒/脈", "R1收/舒/脈", "L2收/舒/脈", "R2收/舒/脈", "信心度"]
    for c, h in enumerate(dup_headers, 1):
        cell = ws_dup.cell(row=4, column=c, value=h)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center")

    row = 5
    for k in sorted(sess_keys.keys()):
        idxs = sess_keys[k]
        if len(idxs) <= 1:
            continue
        for idx in idxs:
            s = sessions[idx]
            ws_dup.cell(row=row, column=1, value=k[0])
            ws_dup.cell(row=row, column=2, value=k[1])
            page_cell = ws_dup.cell(row=row, column=3, value=f"p{s['_page']:02d}")
            png = PAGES_DIR / f"page_{s['_page']:02d}.png"
            if png.exists():
                page_cell.hyperlink = Hyperlink(ref=page_cell.coordinate, target=str(png.absolute()))
                page_cell.font = Font(color="0563C1", underline="single")
            for col_offset, (seq, arm) in enumerate([(1,"L"),(1,"R"),(2,"L"),(2,"R")]):
                r = find_reading(s.get("readings",[]), seq, arm)
                if r:
                    val = f"{r.get('systolic') or '-'}/{r.get('diastolic') or '-'}/{r.get('pulse') or '-'}"
                else:
                    val = "-"
                ws_dup.cell(row=row, column=4+col_offset, value=val)
            ws_dup.cell(row=row, column=8, value=s.get("confidence",""))
            row += 1
        row += 1  # blank row between groups
    for col in "ABCDEFGH":
        ws_dup.column_dimensions[col].width = 14
    ws_dup.column_dimensions["A"].width = 12

    # Page summary sheet
    ws_pg = wb.create_sheet("頁面總覽")
    ws_pg["A1"] = "各頁日期範圍與品質"
    ws_pg["A1"].font = Font(bold=True, size=14)
    pg_headers = ["頁", "日期範圍", "Session 數", "信心度", "備註"]
    for c, h in enumerate(pg_headers, 1):
        cell = ws_pg.cell(row=3, column=c, value=h)
        cell.fill = fill_header
        cell.font = font_header
    page_data = defaultdict(list)
    for s in sessions:
        page_data[s["_page"]].append(s)
    row = 4
    for n in sorted(page_data.keys()):
        ss = page_data[n]
        dates_p = sorted(set(x.get("date","") for x in ss if x.get("date")))
        page_cell = ws_pg.cell(row=row, column=1, value=f"p{n:02d}")
        png = PAGES_DIR / f"page_{n:02d}.png"
        if png.exists():
            page_cell.hyperlink = Hyperlink(ref=page_cell.coordinate, target=str(png.absolute()))
            page_cell.font = Font(color="0563C1", underline="single")
        ws_pg.cell(row=row, column=2, value=f"{dates_p[0]} ~ {dates_p[-1]}" if dates_p else "?")
        ws_pg.cell(row=row, column=3, value=len(ss))
        confs = [x.get("confidence","") for x in ss]
        c_high = sum(1 for c in confs if c=="high")
        c_med = sum(1 for c in confs if c=="medium")
        c_low = sum(1 for c in confs if c=="low")
        conf_cell = ws_pg.cell(row=row, column=4, value=f"H{c_high}/M{c_med}/L{c_low}")
        if c_low > 0 or c_med > c_high:
            conf_cell.fill = fill_med_conf
        else:
            conf_cell.fill = PatternFill("solid", fgColor="C6EFCE")  # green
        # Notes
        n_dup_in_page = sum(1 for s in ss if (s.get("date"), s.get("period")) in [k for k,v in sess_keys.items() if len(v)>1])
        notes = []
        if n_dup_in_page:
            notes.append(f"⚠️ {n_dup_in_page} 筆與其他頁重複")
        if c_low:
            notes.append(f"{c_low} 筆低信心度")
        ws_pg.cell(row=row, column=5, value=" / ".join(notes))
        row += 1
    ws_pg.column_dimensions["A"].width = 6
    ws_pg.column_dimensions["B"].width = 28
    ws_pg.column_dimensions["C"].width = 12
    ws_pg.column_dimensions["D"].width = 16
    ws_pg.column_dimensions["E"].width = 40

    # README sheet
    ws3 = wb.create_sheet("校對說明")
    instructions = [
        ("校對流程說明", True, 14),
        ("", False, 11),
        ("建議從「頁面總覽」開始,先看哪些頁面信心度低或有重複,優先處理那些頁。", True, 11),
        ("", False, 11),
        ("1. 從「頁面總覽」分頁開始", True, 11),
        ("   - 看每頁的信心度分布 (H/M/L) 和重複情況", False, 11),
        ("   - 點「頁」欄超連結打開該頁 PNG", False, 11),
        ("", False, 11),
        ("2. 「重複日期」分頁:處理跨頁重複", True, 11),
        ("   - 同一 (日期, 時段) 出現在多頁 = OCR 把某頁日期讀錯了", False, 11),
        ("   - 比對原圖找出哪頁是錯的 → 在主分頁把錯的列日期改正,並標『已修正』", False, 11),
        ("", False, 11),
        ("3. 「校對檢視」分頁:逐列審查", True, 11),
        ("   - 凍結首列首欄方便捲動", False, 11),
        ("", False, 11),
        ("4. 視覺輔助 (顏色說明)", True, 11),
        ("   紅色整列: 該 (日期+時段) 在多頁出現 (重複)", False, 11),
        ("   橘色儲存格: OCR 標 uncertain 的欄位", False, 11),
        ("   黃色儲存格: 數值超出正常範圍 (S 70-250 / D 40-150 / P 30-200)", False, 11),
        ("   淺紅整列: 該 session 整體信心度 low", False, 11),
        ("   淺黃整列: 該 session 整體信心度 medium", False, 11),
        ("", False, 11),
        ("5. 對照原圖", True, 11),
        ("   - 點「來源頁」欄的 p01 / p02 等超連結 (在 macOS 預覽程式打開)", False, 11),
        ("", False, 11),
        ("6. 修正方式", True, 11),
        ("   - 直接編輯數字儲存格", False, 11),
        ("   - 把「校對狀態」改為「已修正」", False, 11),
        ("   - 沒問題就改為「已確認」", False, 11),
        ("   - 多筆問題太多無法處理 → 「有問題」並寫校對備註", False, 11),
        ("", False, 11),
        ("7. 完成後", True, 11),
        ("   - 通知 Claude:「Phase 1 校對完成」", False, 11),
        ("   - Claude 會把『已確認』+『已修正』狀態的列匯入 SQLite", False, 11),
        ("   - 「未校對」「有問題」狀態的列會跳過", False, 11),
    ]
    for i, (text, bold, size) in enumerate(instructions, 1):
        c = ws3.cell(row=i, column=1, value=text)
        c.font = Font(bold=bold, size=size)
    ws3.column_dimensions["A"].width = 80

    wb.save(OUT_FILE)
    print(f"\nWrote {OUT_FILE}")
    print(f"  rows: {len(sessions)} sessions")


if __name__ == "__main__":
    build()
