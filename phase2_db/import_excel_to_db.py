#!/usr/bin/env python3
"""Import Excel review file into SQLite.

Two modes:
  --unverified   Import all OCR rows as-is, source='ocr_v1'.
                 For dev: get system running before manual review.
                 First deletes all existing ocr_v1 / ocr_v2 rows.
  (default)      Import only 已確認 / 已修正 rows, source='ocr_v2'.
                 First deletes all existing ocr_v1 / ocr_v2 rows.

Manual entries (source='manual' or 'edit') are NEVER deleted.

For duplicate (date, period, seq, arm) rows in the OCR data,
keeps the row with highest confidence (high > medium > low).
"""
import argparse
import sqlite3
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
EXCEL = ROOT / "phase1_ocr" / "output" / "bp_ocr_review.xlsx"
SCHEMA = ROOT / "phase2_db" / "schema.sql"
DB = ROOT / "phase2_db" / "bp.db"

CONF_RANK = {"high": 3, "medium": 2, "low": 1, "": 0, None: 0}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--unverified", action="store_true",
                    help="Import all rows regardless of 校對狀態 (dev mode)")
    ap.add_argument("--db", default=str(DB), help="Output DB path")
    args = ap.parse_args()

    if not EXCEL.exists():
        sys.exit(f"Excel not found: {EXCEL}")

    db_path = Path(args.db)
    init_db = not db_path.exists()
    conn = sqlite3.connect(db_path)
    if init_db:
        conn.executescript(SCHEMA.read_text())
        print(f"Created new DB: {db_path}")
    else:
        # apply schema (idempotent CREATE IF NOT EXISTS)
        conn.executescript(SCHEMA.read_text())

    # Always wipe OCR data and re-import (manual entries preserved)
    cur = conn.cursor()
    cur.execute("DELETE FROM bp_records WHERE source IN ('ocr_v1','ocr_v2')")
    cur.execute("DELETE FROM daily_context WHERE measure_date NOT IN (SELECT DISTINCT measure_date FROM bp_records)")
    n_deleted = cur.rowcount
    print(f"Deleted previous OCR rows (manual entries preserved)")

    wb = load_workbook(EXCEL, data_only=True)
    ws = wb["校對檢視"]

    headers = [c.value for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    source_tag = "ocr_v1" if args.unverified else "ocr_v2"

    # Group rows by (date, period, seq, arm), keep highest confidence
    grouped = defaultdict(list)
    skipped_status = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[col["日期"]]:
            continue
        status = row[col["校對狀態"]] or "未校對"
        if not args.unverified and status not in ("已確認", "已修正"):
            skipped_status += 1
            continue

        page = row[col["來源頁"]]
        date = str(row[col["日期"]])[:10]
        period = row[col["時段"]]
        time_v = row[col["時刻"]]
        if hasattr(time_v, "strftime"):
            time_v = time_v.strftime("%H:%M")
        elif time_v is not None:
            time_v = str(time_v)
        temp = row[col["氣溫°C"]]
        notes = row[col["備註"]] or ""
        conf = row[col["信心度"]] or ""

        for seq, arm in [(1, "L"), (1, "R"), (2, "L"), (2, "R")]:
            sys_v = row[col[f"{arm}{seq}收"]]
            dia_v = row[col[f"{arm}{seq}舒"]]
            pul_v = row[col[f"{arm}{seq}脈"]]
            if sys_v is None and dia_v is None and pul_v is None:
                continue
            grouped[(date, period, seq, arm)].append({
                "time": time_v, "temp": temp, "notes": notes,
                "sys": sys_v, "dia": dia_v, "pul": pul_v,
                "page": page, "conf": conf,
            })

    # For each unique (date, period, seq, arm), pick highest confidence
    n_imported = 0
    n_dup_resolved = 0
    daily_temp = {}

    for key, candidates in grouped.items():
        if len(candidates) > 1:
            n_dup_resolved += 1
            candidates.sort(key=lambda c: -CONF_RANK.get(c["conf"], 0))
        c = candidates[0]
        date, period, seq, arm = key

        if c["temp"] is not None:
            daily_temp[date] = c["temp"]

        try:
            cur.execute(
                "INSERT OR REPLACE INTO bp_records "
                "(user_id, measure_date, period, measure_time, sequence, arm, "
                " systolic, diastolic, pulse, notes, source, source_ref) "
                "VALUES (1, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (date, period, c["time"], seq, arm,
                 c["sys"], c["dia"], c["pul"], c["notes"], source_tag, c["page"]),
            )
            n_imported += 1
        except sqlite3.IntegrityError as e:
            print(f"  WARN: skipping {key}: {e}")

    # Daily context (temperature)
    n_temp = 0
    for date, temp in daily_temp.items():
        cur.execute(
            "INSERT INTO daily_context(user_id, measure_date, temperature_c) "
            "VALUES (1, ?, ?) "
            "ON CONFLICT(user_id, measure_date) DO UPDATE SET temperature_c = excluded.temperature_c",
            (date, float(temp)),
        )
        n_temp += 1

    conn.commit()

    # Summary
    print(f"\n=== Import Summary ===")
    print(f"Mode: {'unverified (dev)' if args.unverified else 'verified'}")
    print(f"Source tag: {source_tag}")
    print(f"Imported readings: {n_imported}")
    print(f"Daily context (temperature): {n_temp}")
    print(f"Skipped (not 已確認/已修正): {skipped_status}")
    print(f"Duplicate slots resolved by confidence: {n_dup_resolved}")

    n_days = cur.execute("SELECT COUNT(DISTINCT measure_date) FROM bp_records").fetchone()[0]
    n_total = cur.execute("SELECT COUNT(*) FROM bp_records").fetchone()[0]
    n_manual = cur.execute("SELECT COUNT(*) FROM bp_records WHERE source IN ('manual','edit')").fetchone()[0]
    print(f"\nDB stats:")
    print(f"  Total BP rows: {n_total}")
    print(f"  Unique days: {n_days}")
    print(f"  OCR rows: {n_total - n_manual}")
    print(f"  Manual rows: {n_manual}")
    print(f"  DB: {db_path}")

    conn.close()


if __name__ == "__main__":
    main()
